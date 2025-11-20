"""
Analysis modules for document, code, and on-chain data
"""

import os
import re
import requests
from typing import Dict, List
from datetime import datetime, timedelta

class DocumentAnalyzer:
    """Analyzes pitch decks, whitepapers, and other documents"""
    
    def analyze_pdf(self, filepath: str) -> Dict:
        """
        Analyze a PDF document (pitch deck or whitepaper)
        
        Returns quality score based on:
        - Document length
        - Structure
        - Content quality
        - Technical depth
        """
        try:
            import PyPDF2
            
            with open(filepath, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                num_pages = len(pdf_reader.pages)
                
                # Extract text from all pages
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
            
            # Analyze content
            word_count = len(text.split())
            
            # Check for key sections
            has_problem = bool(re.search(r'problem|challenge|pain point', text, re.IGNORECASE))
            has_solution = bool(re.search(r'solution|approach|our product', text, re.IGNORECASE))
            has_market = bool(re.search(r'market|tam|addressable market', text, re.IGNORECASE))
            has_team = bool(re.search(r'team|founder|advisor', text, re.IGNORECASE))
            has_traction = bool(re.search(r'traction|milestone|achievement', text, re.IGNORECASE))
            has_financials = bool(re.search(r'revenue|financial|projection', text, re.IGNORECASE))
            
            # Calculate score
            score = 0
            
            # Page count (max 20 points)
            if num_pages >= 10:
                score += 20
            elif num_pages >= 5:
                score += 15
            else:
                score += 10
            
            # Word count (max 20 points)
            if word_count >= 2000:
                score += 20
            elif word_count >= 1000:
                score += 15
            else:
                score += 10
            
            # Key sections (10 points each, max 60 points)
            section_score = sum([
                has_problem, has_solution, has_market,
                has_team, has_traction, has_financials
            ]) * 10
            score += section_score
            
            return {
                'quality_score': min(score, 100),
                'num_pages': num_pages,
                'word_count': word_count,
                'has_key_sections': {
                    'problem': has_problem,
                    'solution': has_solution,
                    'market': has_market,
                    'team': has_team,
                    'traction': has_traction,
                    'financials': has_financials
                },
                'analysis_type': 'pdf'
            }
            
        except Exception as e:
            print(f"⚠️  PDF analysis failed: {e}")
            return {
                'quality_score': 50,
                'error': str(e),
                'analysis_type': 'pdf'
            }
    
    def analyze_video(self, filepath: str) -> Dict:
        """
        Analyze a demo video
        
        Returns quality score based on:
        - Video length
        - File size (indicates quality)
        """
        try:
            file_size = os.path.getsize(filepath)
            
            # Estimate quality based on file size
            # Larger files generally indicate better quality
            score = 0
            
            if file_size > 50 * 1024 * 1024:  # > 50MB
                score = 85
            elif file_size > 20 * 1024 * 1024:  # > 20MB
                score = 70
            elif file_size > 5 * 1024 * 1024:  # > 5MB
                score = 55
            else:
                score = 40
            
            return {
                'quality_score': score,
                'file_size_mb': file_size / (1024 * 1024),
                'analysis_type': 'video'
            }
            
        except Exception as e:
            print(f"⚠️  Video analysis failed: {e}")
            return {
                'quality_score': 50,
                'error': str(e),
                'analysis_type': 'video'
            }
    
    def analyze_spreadsheet(self, filepath: str) -> Dict:
        """
        Analyze financial spreadsheet
        
        Returns quality score based on:
        - Number of sheets
        - Data completeness
        """
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(filepath)
            num_sheets = len(wb.sheetnames)
            
            score = 0
            
            # More sheets generally indicate more detailed financials
            if num_sheets >= 5:
                score = 90
            elif num_sheets >= 3:
                score = 75
            elif num_sheets >= 2:
                score = 60
            else:
                score = 45
            
            return {
                'quality_score': score,
                'num_sheets': num_sheets,
                'sheet_names': wb.sheetnames,
                'analysis_type': 'spreadsheet'
            }
            
        except Exception as e:
            print(f"⚠️  Spreadsheet analysis failed: {e}")
            return {
                'quality_score': 50,
                'error': str(e),
                'analysis_type': 'spreadsheet'
            }


class CodeAnalyzer:
    """Analyzes GitHub repositories for code quality"""
    
    def __init__(self):
        self.github_token = os.getenv('GITHUB_TOKEN')
        self.headers = {}
        if self.github_token:
            self.headers['Authorization'] = f'token {self.github_token}'
    
    def analyze_repository(self, repo_name: str) -> Dict:
        """
        Analyze a GitHub repository
        
        Args:
            repo_name: Format "owner/repo"
            
        Returns quality score based on:
        - Commit frequency
        - Number of contributors
        - Code quality metrics
        - Documentation
        """
        try:
            base_url = f"https://api.github.com/repos/{repo_name}"
            
            # Get repository info
            repo_response = requests.get(base_url, headers=self.headers)
            repo_data = repo_response.json()
            
            # Get commit activity
            commits_url = f"{base_url}/commits"
            commits_response = requests.get(
                commits_url,
                headers=self.headers,
                params={'per_page': 100}
            )
            commits = commits_response.json() if isinstance(commits_response.json(), list) else []
            
            # Get contributors
            contributors_url = f"{base_url}/contributors"
            contributors_response = requests.get(contributors_url, headers=self.headers)
            contributors = contributors_response.json() if isinstance(contributors_response.json(), list) else []
            
            # Calculate metrics
            stars = repo_data.get('stargazers_count', 0)
            forks = repo_data.get('forks_count', 0)
            open_issues = repo_data.get('open_issues_count', 0)
            has_readme = repo_data.get('has_readme', False)
            has_license = repo_data.get('license') is not None
            
            num_commits = len(commits)
            num_contributors = len(contributors)
            
            # Calculate recent activity (last 90 days)
            recent_commits = 0
            cutoff_date = datetime.utcnow() - timedelta(days=90)
            
            for commit in commits:
                commit_date_str = commit.get('commit', {}).get('author', {}).get('date', '')
                if commit_date_str:
                    commit_date = datetime.strptime(commit_date_str, '%Y-%m-%dT%H:%M:%SZ')
                    if commit_date > cutoff_date:
                        recent_commits += 1
            
            # Calculate score
            score = 0
            
            # Stars (max 15 points)
            if stars >= 100:
                score += 15
            elif stars >= 50:
                score += 12
            elif stars >= 10:
                score += 8
            else:
                score += 5
            
            # Recent activity (max 25 points)
            if recent_commits >= 50:
                score += 25
            elif recent_commits >= 20:
                score += 20
            elif recent_commits >= 10:
                score += 15
            else:
                score += 10
            
            # Contributors (max 20 points)
            if num_contributors >= 10:
                score += 20
            elif num_contributors >= 5:
                score += 15
            elif num_contributors >= 2:
                score += 10
            else:
                score += 5
            
            # Documentation (max 20 points)
            doc_score = 0
            if has_readme:
                doc_score += 10
            if has_license:
                doc_score += 10
            score += doc_score
            
            # Code health (max 20 points)
            health_score = 20
            if open_issues > 50:
                health_score -= 10
            elif open_issues > 20:
                health_score -= 5
            score += health_score
            
            return {
                'score': min(score, 100),
                'stars': stars,
                'forks': forks,
                'commits': num_commits,
                'recent_commits_90d': recent_commits,
                'contributors': num_contributors,
                'open_issues': open_issues,
                'has_readme': has_readme,
                'has_license': has_license,
                'analysis_type': 'github'
            }
            
        except Exception as e:
            print(f"⚠️  GitHub analysis failed: {e}")
            return {
                'score': 50,
                'error': str(e),
                'analysis_type': 'github'
            }


class OnChainAnalyzer:
    """Analyzes on-chain activity on Sui"""
    
    def __init__(self):
        self.sui_rpc = os.getenv('SUI_RPC', 'https://fullnode.testnet.sui.io:443')
    
    def analyze_wallet(self, address: str) -> Dict:
        """
        Analyze wallet activity on Sui
        
        Returns score based on:
        - Transaction count
        - Balance
        - Contract interactions
        - Age of wallet
        """
        try:
            # Get transactions
            response = requests.post(
                self.sui_rpc,
                json={
                    "jsonrpc": "2.0",
                    "id": 1,
                    "method": "suix_getTransactions",
                    "params": [
                        address,
                        None,
                        100,
                        False
                    ]
                }
            )
            
            data = response.json()
            
            # Get balance
            balance_response = requests.post(
                self.sui_rpc,
                json={
                    "jsonrpc": "2.0",
                    "id": 1,
                    "method": "suix_getBalance",
                    "params": [address, "0x2::sui::SUI"]
                }
            )
            
            balance_data = balance_response.json()
            balance = int(balance_data.get('result', {}).get('totalBalance', 0)) / 1e9  # Convert to SUI
            
            # Calculate metrics
            tx_count = len(data.get('result', {}).get('data', []))
            
            # Calculate score
            score = 0
            
            # Transaction count (max 40 points)
            if tx_count >= 100:
                score += 40
            elif tx_count >= 50:
                score += 30
            elif tx_count >= 20:
                score += 20
            elif tx_count >= 5:
                score += 10
            else:
                score += 5
            
            # Balance (max 30 points)
            if balance >= 1000:
                score += 30
            elif balance >= 100:
                score += 25
            elif balance >= 10:
                score += 20
            elif balance >= 1:
                score += 15
            else:
                score += 5
            
            # Activity score (max 30 points)
            # Active wallets get higher scores
            if tx_count > 0:
                score += 30
            
            return {
                'score': min(score, 100),
                'transaction_count': tx_count,
                'balance_sui': balance,
                'analysis_type': 'on_chain'
            }
            
        except Exception as e:
            print(f"⚠️  On-chain analysis failed: {e}")
            return {
                'score': 50,
                'error': str(e),
                'analysis_type': 'on_chain'
            }
